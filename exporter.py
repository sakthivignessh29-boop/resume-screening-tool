import io
import os
from datetime import datetime

def export_excel(results, job_title="Job Analysis"):
    """
    Export analysis results to an Excel workbook.
    Returns BytesIO buffer.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    # ── Color palette ──────────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="1E1B4B")
    HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    TITLE_FONT    = Font(name="Calibri", bold=True, size=14, color="1E1B4B")
    ALT_FILL      = PatternFill("solid", fgColor="F0F0FF")
    CENTER        = Alignment(horizontal="center", vertical="center")
    LEFT          = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    BORDER_SIDE   = Side(style="thin", color="CCCCCC")
    THIN_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                           top=BORDER_SIDE, bottom=BORDER_SIDE)

    # ── Title Row ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"] = f"AI Resume Screening — {job_title}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(color="666666", italic=True, size=9)
    ws["A2"].alignment = CENTER

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 14

    # ── Headers ────────────────────────────────────────────────────────────────
    headers = [
        "Rank", "Candidate Name", "Filename", "Total Score (%)",
        "Keyword Score", "Experience Score", "Education Score",
        "Years Exp.", "Education Level"
    ]
    col_widths = [6, 24, 28, 16, 16, 17, 17, 12, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 18

    # ── Data Rows ──────────────────────────────────────────────────────────────
    sorted_results = sorted(results, key=lambda x: x.get('total_score', 0), reverse=True)

    for row_idx, res in enumerate(sorted_results, start=4):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        rank = row_idx - 3

        values = [
            rank,
            res.get('candidate_name', 'Unknown'),
            res.get('filename', ''),
            round(res.get('total_score', 0), 1),
            round(res.get('keyword_score', 0), 1),
            round(res.get('experience_score', 0), 1),
            round(res.get('education_score', 0), 1),
            res.get('years_experience', 0),
            res.get('education_level', 'N/A'),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CENTER if col_idx in (1, 4, 5, 6, 7, 8) else LEFT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 16

    # ── Keywords Sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Keyword Details")
    kw_headers = ["Candidate", "Keyword", "Matched", "Weight", "Occurrences"]
    kw_widths   = [28, 24, 12, 10, 14]
    for col_idx, (h, w) in enumerate(zip(kw_headers, kw_widths), start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    row = 2
    for res in sorted_results:
        name = res.get('candidate_name', res.get('filename', 'Unknown'))
        all_kws = res.get('matched_keywords', []) + res.get('missing_keywords', [])
        for kw in all_kws:
            ws2.cell(row=row, column=1, value=name)
            ws2.cell(row=row, column=2, value=kw.get('keyword', ''))
            matched_cell = ws2.cell(row=row, column=3, value='✓' if kw.get('matched') else '✗')
            matched_cell.font = Font(
                color="00AA00" if kw.get('matched') else "CC0000", bold=True)
            ws2.cell(row=row, column=4, value=kw.get('weight', 1.0))
            ws2.cell(row=row, column=5, value=kw.get('occurrences', 0))
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def export_pdf(results, job_title="Job Analysis"):
    """
    Export analysis results to a styled PDF.
    Returns BytesIO buffer.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer, HRFlowable)
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise ImportError("reportlab is required for PDF export. Run: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        title=f"Resume Screening — {job_title}"
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                 fontSize=18, textColor=colors.HexColor('#1E1B4B'),
                                 alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle('Sub', parent=styles['Normal'],
                                 fontSize=9, textColor=colors.grey,
                                 alignment=TA_CENTER, spaceAfter=12)
    body_style  = ParagraphStyle('Body', parent=styles['Normal'],
                                 fontSize=8, textColor=colors.HexColor('#111111'))

    story = []
    story.append(Paragraph(f"AI Resume Screening Report", title_style))
    story.append(Paragraph(f"Job: {job_title} &nbsp;|&nbsp; Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", sub_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#6C63FF'), spaceAfter=12))

    # ── Results table ──────────────────────────────────────────────────────────
    sorted_results = sorted(results, key=lambda x: x.get('total_score', 0), reverse=True)

    table_data = [["Rank", "Candidate Name", "Filename",
                   "Total\nScore", "Keyword\nScore", "Exp.\nScore",
                   "Edu.\nScore", "Years\nExp.", "Education"]]

    for rank, res in enumerate(sorted_results, start=1):
        matched_count = len(res.get('matched_keywords', []))
        total_kw = matched_count + len(res.get('missing_keywords', []))
        table_data.append([
            str(rank),
            res.get('candidate_name', 'Unknown'),
            res.get('filename', ''),
            f"{res.get('total_score', 0):.1f}",
            f"{res.get('keyword_score', 0):.1f}",
            f"{res.get('experience_score', 0):.1f}",
            f"{res.get('education_score', 0):.1f}",
            str(res.get('years_experience', 0)),
            res.get('education_level', 'N/A'),
        ])

    col_widths = [1.2*cm, 4*cm, 5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 1.8*cm, 3*cm]

    HEADER_DARK  = colors.HexColor('#1E1B4B')
    ACCENT       = colors.HexColor('#6C63FF')
    ALT          = colors.HexColor('#F0F0FF')
    WHITE        = colors.white
    GOLD         = colors.HexColor('#FFD700')
    SILVER       = colors.HexColor('#C0C0C0')
    BRONZE       = colors.HexColor('#CD7F32')

    ts = TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0), HEADER_DARK),
        ('TEXTCOLOR',   (0, 0), (-1, 0), WHITE),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0), 8),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, ALT]),
        ('FONTSIZE',    (0, 1), (-1, -1), 8),
        ('GRID',        (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWHEIGHT',   (0, 0), (-1, -1), 22),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
    ])
    # Medal colors for top 3
    for medal_row, medal_color in [(1, GOLD), (2, SILVER), (3, BRONZE)]:
        if medal_row < len(table_data):
            ts.add('BACKGROUND', (0, medal_row), (0, medal_row), medal_color)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(ts)
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # ── Per-candidate breakdown ────────────────────────────────────────────────
    story.append(Paragraph("Score Breakdown by Candidate", ParagraphStyle(
        'SectionTitle', fontSize=12, textColor=HEADER_DARK,
        fontName='Helvetica-Bold', spaceBefore=8, spaceAfter=4)))
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=6))

    for rank, res in enumerate(sorted_results, start=1):
        name = res.get('candidate_name', res.get('filename', 'Unknown'))
        matched = [k['keyword'] for k in res.get('matched_keywords', [])]
        missing = [k['keyword'] for k in res.get('missing_keywords', [])]
        breakdown = res.get('score_breakdown', {})

        story.append(Paragraph(
            f"<b>#{rank} {name}</b> — Total Score: <b>{res.get('total_score', 0):.1f}/100</b>",
            ParagraphStyle('CandName', fontSize=9, textColor=HEADER_DARK, spaceBefore=6)))

        lines = []
        if breakdown.get('keyword'):
            lines.append(f"• Keyword Score: {breakdown['keyword'].get('raw',0):.1f} × {breakdown['keyword'].get('weight',0):.0%} = {breakdown['keyword'].get('contribution',0):.1f} pts")
        if breakdown.get('experience'):
            lines.append(f"• Experience Score: {breakdown['experience'].get('raw',0):.1f} × {breakdown['experience'].get('weight',0):.0%} = {breakdown['experience'].get('contribution',0):.1f} pts ({breakdown['experience'].get('years',0)} yrs)")
        if breakdown.get('education'):
            lines.append(f"• Education Score: {breakdown['education'].get('raw',0):.1f} × {breakdown['education'].get('weight',0):.0%} = {breakdown['education'].get('contribution',0):.1f} pts ({breakdown['education'].get('level','N/A')})")
        if matched:
            lines.append(f"• Matched Keywords: {', '.join(matched[:10])}")
        if missing:
            lines.append(f"• Missing Keywords: {', '.join(missing[:10])}")

        for line in lines:
            story.append(Paragraph(line, ParagraphStyle('BreakLine', fontSize=7.5,
                textColor=colors.HexColor('#333333'), leftIndent=12, spaceAfter=1)))

    doc.build(story)
    buf.seek(0)
    return buf
